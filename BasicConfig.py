import aiohttp
import asyncio
import openpyxl
from bs4 import BeautifulSoup

# Laad het Excel-bestand
FILE_NAME = "wielermanager.xlsx"

# Haal renners uit Excel
def get_riders():
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb["Team"]
    return [row[0].value for row in ws.iter_rows(min_row=2, max_col=1) if row[0].value]

# Haal wedstrijden uit Excel
def get_races():
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb["Wedstrijden"]
    return [row[0].value for row in ws.iter_rows(min_row=2, max_col=1) if row[0].value]

# Scrape startlijst van ProCyclingStats
async def get_startlist(session, race_name):
    race_url = f"https://www.procyclingstats.com/race/{race_name.replace(' ', '-').lower()}/2025/startlist"
    print(f"Scraping URL: {race_url}")  # Debugging: Controleer of de URL correct is

    async with session.get(race_url) as response:
        if response.status != 200:
            print(f"Fout bij ophalen van {race_name} ({response.status})")
            return []
        
        soup = BeautifulSoup(await response.text(), "html.parser")

        # Nieuwe correcte selector op basis van de HTML-structuur
        startlist = [rider.text.strip() for rider in soup.select("div.ridersCont ul li a")]

        if not startlist:
            print("⚠️ Geen renners gevonden! Mogelijk een probleem met de HTML-structuur.")
        
        print(f"Startlijst voor {race_name}: {startlist}")  # Debugging: Print de rennerslijst
        return startlist

# Hoofdfunctie: vergelijkt renners met startlijst en update Excel
async def update_excel():
    riders = get_riders()
    races = get_races()
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb["Wedstrijden"]

    async with aiohttp.ClientSession() as session:
        for i, race in enumerate(races, start=2):
            startlist = await get_startlist(session, race)
            print(f"Startlijst voor {race}:", startlist)  # Debug-print
            count = sum(1 for rider in riders if any(rider.lower() in s.lower() or s.lower() in rider.lower() for s in startlist))
            ws.cell(row=i, column=2, value=count)  # Kolom B updaten

    wb.save(FILE_NAME)

# Debugging: check of data goed wordt gelezen
print("Gelezen renners:", get_riders())
print("Gelezen wedstrijden:", get_races())

# Voer het script uit
if __name__ == "__main__":
    asyncio.run(update_excel())